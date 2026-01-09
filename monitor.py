import instaloader
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import logging

ARQUIVO_EXCEL = r"C:\Users\joliv\OneDrive\Desktop\projetos\monitor_instagram.xlsx"
PERFIL = "jones.manoel"

# função para criar log

logging.basicConfig( filename=r"C:\Users\joliv\OneDrive\Desktop\projetos\monitor.log", level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s", )

def log(msg):
    print(msg) # aparece no console
    logging.info(msg) # grava no arquivo

# função para abrir sessão.

def get_loader():
    L = instaloader.Instaloader()
    L.load_session_from_file("sojjoel")
    return L


# --- 1. Carregar Excel e abas ---

def carregar_planilha():
    wb = load_workbook(ARQUIVO_EXCEL)
    posts_ws = wb["posts"]
    metricas_ws = wb["metricas"]
    return wb, posts_ws, metricas_ws

# --- 2. Ler posts já cadastrados (shortcode -> id_post) ---



def mapear_posts_existentes(posts_ws):
    mapa = {}
    ultima_linha = posts_ws.max_row
    for row in range(2, ultima_linha + 1):
        shortcode = posts_ws[f"B{row}"].value
        id_post = posts_ws[f"A{row}"].value
        ativo = posts_ws[f"H{row}"].value
        mapa[shortcode] = {"id_post": id_post, "linha": row, "ativo": ativo}
    return mapa

# --- 3. Obter próximo id_post disponível ---

def proximo_id_post(posts_ws):
    ultima_linha = posts_ws.max_row
    if ultima_linha < 2:
        return 1
    ultimo_id = posts_ws[f"A{ultima_linha}"].value
    return int(ultimo_id) + 1

#--- 4. Conectar no Instaloader e pegar posts recentes ---


def obter_posts_recentes(L, perfil, limite=5):
    profile = instaloader.Profile.from_username(L.context, perfil)
    posts = profile.get_posts() # iterador de posts
    lista = []
    for i, post in enumerate(posts):
        if i >= limite: break
        lista.append(post)
    return lista

# --- 5. Adicionar novos posts na aba POSTS ---


def registrar_novos_posts(posts_ws, mapa_posts, posts_recentes):
    mudou = False
    prox_id = proximo_id_post(posts_ws)
    for post in posts_recentes:
        shortcode = post.shortcode
        if shortcode in mapa_posts:
            continue # já existe
        data_pub = post.date_local # datetime
        tipo = post.typename # tipo de conteúdo (GraphImage, GraphVideo, etc.)
        url = f"https://www.instagram.com/p/{shortcode}/"
        linha = posts_ws.max_row + 1
        posts_ws[f"A{linha}"] = prox_id
        posts_ws[f"B{linha}"] = shortcode
        posts_ws[f"C{linha}"] = PERFIL
        posts_ws[f"D{linha}"] = data_pub.date()
        posts_ws[f"E{linha}"] = data_pub.time().replace(microsecond=0)
        posts_ws[f"F{linha}"] = tipo
        posts_ws[f"G{linha}"] = url
        posts_ws[f"H{linha}"] = "SIM" # ativo
        mapa_posts[shortcode] = {"id_post": prox_id, "linha": linha, "ativo": "SIM"}
        log(f"Novo post detectado: {shortcode} (ID {prox_id})")
        prox_id += 1
        mudou = True
        
    return mudou

# --- 6. Atualizar métricas dos posts ativos ---

def atualizar_metricas(L, metricas_ws, posts_ws, mapa_posts):
    agora = datetime.now()
    data_coleta = agora.date()
    hora_coleta = agora.time().replace(microsecond=0)
    for shortcode, info in mapa_posts.items():
        if info["ativo"] != "SIM":
            continue
        try:
            post = instaloader.Post.from_shortcode(L.context, shortcode) #

        except Exception as e:
            log(f"Erro ao carregar post {shortcode}: {e}")
            continue
        log(f"Atualizando métricas do post {shortcode} (ID {info['id_post']})")
        
        curtidas = post.likes # número de likes
        comentarios = post.comments # total de comentários
        linha = metricas_ws.max_row + 1
        
        metricas_ws[f"A{linha}"] = info["id_post"]
        metricas_ws[f"B{linha}"] = data_coleta
        metricas_ws[f"C{linha}"] = hora_coleta
        metricas_ws[f"D{linha}"] = curtidas
        metricas_ws[f"E{linha}"] = comentarios
        metricas_ws[f"F{linha}"] = (data_coleta - post.date_local.date()).days


# --- 7. Verificar expiração (6 meses) ---


def verificar_expiracao(posts_ws, mapa_posts, dias_limite=180):
    hoje = datetime.now().date()
    for shortcode, info in mapa_posts.items():
        
        if info["ativo"] != "SIM":
            continue
        
        linha = info["linha"]
        data_pub = posts_ws[f"D{linha}"].value # data_publicacao
        if not isinstance(data_pub, datetime):
            # se vier como date puro, ok
            data_pub_date = data_pub
        else:
            data_pub_date = data_pub.date()
            
        idade = (hoje - data_pub_date).days
        if idade >= dias_limite:
            posts_ws[f"H{linha}"] = "NÃO"
            info["ativo"] = "NÃO"
            log(f"Post {shortcode} expirou com {idade} dias. Marcado como inativo.")

# --- 8. Função principal do ciclo ---


def ciclo_monitoramento():
    
    log("Iniciando ciclo de monitoramento...")
    
    L = get_loader() # <<--- loader logado
    wb, posts_ws, metricas_ws = carregar_planilha()
    mapa_posts = mapear_posts_existentes(posts_ws)

    # 1) Buscar posts recentes e registrar novos

    log(f"Buscando posts recentes do perfil {PERFIL}...")
    
    posts_recentes = obter_posts_recentes(L, PERFIL, limite=10)
    novos = registrar_novos_posts(posts_ws, mapa_posts, posts_recentes)
    
    if novos: print("Novos posts adicionados na aba POSTS.")

    # 2) Atualizar métricas dos posts ativos

    atualizar_metricas(L, metricas_ws, posts_ws, mapa_posts)
    print("Métricas atualizadas.")

    # 3) Verificar expiração (7 dias)

    verificar_expiracao(posts_ws, mapa_posts, dias_limite=7)
    print("Expiração verificada.")

    # 4) Salvar planilha

    wb.save(ARQUIVO_EXCEL)
    print("Planilha salva.")

    log("Ciclo concluído e planilha salva.")


if __name__ == "__main__":
    ciclo_monitoramento()















































































    
